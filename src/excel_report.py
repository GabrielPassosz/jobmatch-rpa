from pathlib import Path

import pandas as pd
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)
from openpyxl.utils import get_column_letter


# ==========================================
# SALVAR RELATÓRIO EXCEL
# ==========================================

def salvar_relatorio_excel(
    resultados,
    pasta_saida,
):
    pasta_saida = Path(
        pasta_saida
    )

    pasta_saida.mkdir(
        parents=True,
        exist_ok=True
    )

    caminho_relatorio = (
        pasta_saida
        / "jobmatch_resultado.xlsx"
    )

    # ==============================
    # DATAFRAME PRINCIPAL
    # ==============================

    df = pd.DataFrame(
        resultados
    )

    # ==============================
    # DATAFRAME RESUMO
    # ==============================

    resumo = pd.DataFrame({
        "Indicador": [
            "Total de vagas",
            "Vagas para aplicar",
            "Vagas para avaliar",
            "Baixa compatibilidade",
            "Maior score",
            "Score médio",
        ],

        "Resultado": [
            len(df),

            len(
                df[
                    df["Recomendacao"]
                    == "APLICAR"
                ]
            ),

            len(
                df[
                    df["Recomendacao"]
                    == "AVALIAR"
                ]
            ),

            len(
                df[
                    df["Recomendacao"]
                    == "BAIXA COMPATIBILIDADE"
                ]
            ),

            (
                df["Score"].max()
                if not df.empty
                else 0
            ),

            (
                round(
                    df["Score"].mean(),
                    2
                )
                if not df.empty
                else 0
            ),
        ]
    })

    # ==============================
    # CRIAR EXCEL
    # ==============================

    with pd.ExcelWriter(
        caminho_relatorio,
        engine="openpyxl"
    ) as writer:

        # ==============================
        # ABA RANKING
        # ==============================

        df.to_excel(
            writer,
            sheet_name="Ranking",
            index=False
        )

        # ==============================
        # ABA RESUMO
        # ==============================

        resumo.to_excel(
            writer,
            sheet_name="Resumo",
            index=False
        )

        workbook = writer.book

        # ==================================
        # FORMATAR ABA RANKING
        # ==================================

        worksheet = workbook["Ranking"]

        formatar_cabecalho(
            worksheet
        )

        ajustar_largura_colunas(
            worksheet
        )

        aplicar_cores_recomendacao(
            worksheet
        )

        # ==================================
        # FORMATAR ABA RESUMO
        # ==================================

        worksheet_resumo = (
            workbook["Resumo"]
        )

        formatar_cabecalho(
            worksheet_resumo
        )

        ajustar_largura_colunas(
            worksheet_resumo
        )

    return caminho_relatorio


# ==========================================
# FORMATAR CABEÇALHO
# ==========================================

def formatar_cabecalho(
    worksheet
):
    fill = PatternFill(
        start_color="1E3A8A",
        end_color="1E3A8A",
        fill_type="solid"
    )

    font = Font(
        color="FFFFFF",
        bold=True
    )

    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for cell in worksheet[1]:

        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


# ==========================================
# AJUSTAR LARGURA
# ==========================================

def ajustar_largura_colunas(
    worksheet
):
    for column in worksheet.columns:

        tamanho_maximo = 0

        letra_coluna = (
            get_column_letter(
                column[0].column
            )
        )

        for cell in column:

            try:

                if len(str(cell.value)) > tamanho_maximo:

                    tamanho_maximo = len(
                        str(cell.value)
                    )

            except:
                pass

        largura_ajustada = (
            tamanho_maximo + 5
        )

        worksheet.column_dimensions[
            letra_coluna
        ].width = largura_ajustada


# ==========================================
# COLORIR RECOMENDAÇÃO
# ==========================================

def aplicar_cores_recomendacao(
    worksheet
):
    for row in worksheet.iter_rows(
        min_row=2
    ):

        recomendacao = (
            row[6].value
        )

        if recomendacao == "APLICAR":

            fill = PatternFill(
                start_color="DCFCE7",
                end_color="DCFCE7",
                fill_type="solid"
            )

        elif recomendacao == "AVALIAR":

            fill = PatternFill(
                start_color="FEF3C7",
                end_color="FEF3C7",
                fill_type="solid"
            )

        else:

            fill = PatternFill(
                start_color="FEE2E2",
                end_color="FEE2E2",
                fill_type="solid"
            )

        for cell in row:

            cell.fill = fill